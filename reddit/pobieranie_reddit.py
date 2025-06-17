import praw
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# Konfiguracja klienta Reddit API
reddit = praw.Reddit(
    client_id="twhA-1yWd6CFHqqBIRyg7A",
    client_secret="5Yr6ygmmSJJYXYOoGsuJERyk0cgcAQ",
    user_agent="script:analiza-sondazy-wyborczych:v1.0 (by /ActionVast9655)"
)

# Parametry wyszukiwania
subreddit_name = "Polska"
search_term = "biejat"
dni_wstecz = 45  # od początku maja do dziś
limit = 1000  # możesz zwiększyć do 2000–5000, ale PRAW zwróci mniej, jeśli nie ma więcej

def pobierz_posty(subreddit_name, search_term, limit, dni_wstecz):
    subreddit = reddit.subreddit(subreddit_name)
    posts = []
    cutoff_timestamp = (datetime.now() - timedelta(days=dni_wstecz)).timestamp()

    for post in subreddit.search(search_term, limit=limit):
        if post.created_utc < cutoff_timestamp:
            continue

        posts.append([
            datetime.fromtimestamp(post.created_utc).strftime('%Y-%m-%d %H:%M:%S'),
            post.author.name if post.author else "[usunięty]",
            post.title,
            post.selftext,
            post.score,
            post.num_comments,
            f"https://www.reddit.com{post.permalink}"
        ])

    return posts

print(f"Szukam postów zawierających '{search_term}' z ostatnich {dni_wstecz} dni w r/{subreddit_name}...")

posts = pobierz_posty(subreddit_name, search_term, limit, dni_wstecz)

if not posts:
    print("Brak wyników.")
    exit()

df = pd.DataFrame(posts, columns=["Data", "Użytkownik", "Tytuł", "Treść", "Punkty", "Komentarze", "Link"])
df['Data'] = pd.to_datetime(df['Data'])
df = df.sort_values(by='Data', ascending=False)

csv_filename = f"reddit_{search_term}_{subreddit_name}.csv"
excel_filename = f"reddit_{search_term}_{subreddit_name}.xlsx"

df.to_csv(csv_filename, index=False, encoding='utf-8')

try:
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Posty Reddit')

        ws = writer.sheets['Posty Reddit']

        # Stylizacja: pogrubienie nagłówków, zawijanie tekstu, kolory
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="4F81BD")

        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = fill
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                15 if column in ["Data", "Użytkownik", "Punkty", "Komentarze"]
                else 45 if column == "Link"
                else 70 if column == "Treść"
                else 40
            )

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    print(f"\n✅ Zapisano dane do:")
    print(f"📄 CSV: {csv_filename}")
    print(f"📘 Excel: {excel_filename}")

except Exception as e:
    print(f"❌ Excel nie został zapisany: {e}")
    print(f"📄 Zapisano tylko CSV: {csv_filename}")

# Podsumowanie
print("\n📊 Podsumowanie:")
print(f"🔹 Liczba postów: {len(df)}")
print(f"📅 Najstarszy post: {df['Data'].min().strftime('%Y-%m-%d')}")
print(f"📅 Najnowszy post: {df['Data'].max().strftime('%Y-%m-%d')}")
print(f"⬆️ Najwięcej punktów: {df['Punkty'].max()} ({df.loc[df['Punkty'].idxmax()]['Tytuł']})")
print(f"💬 Najwięcej komentarzy: {df['Komentarze'].max()} ({df.loc[df['Komentarze'].idxmax()]['Tytuł']})")
