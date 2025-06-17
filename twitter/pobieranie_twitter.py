import snscrape.modules.twitter as sntwitter
import pandas as pd
from datetime import datetime, timedelta

# Parametry wyszukiwania
search_term = "sondaż"
dni_wstecz = 45  # od początku maja do dziś
max_tweets = 1000  # limit tweetów do pobrania

# Oblicz datę początkową
date_since = (datetime.now() - timedelta(days=dni_wstecz)).strftime('%Y-%m-%d')


def pobierz_tweety(search_term, max_tweets, date_since):
    tweets = []
    query = f"{search_term} since:{date_since}"

    for i, tweet in enumerate(sntwitter.TwitterSearchScraper(query).get_items()):
        if i >= max_tweets:
            break

        tweets.append([
            tweet.date.strftime('%Y-%m-%d %H:%M:%S'),
            tweet.user.username,
            tweet.content,
            tweet.likeCount,
            tweet.retweetCount,
            f"https://twitter.com/{tweet.user.username}/status/{tweet.id}"
        ])
    return tweets


print(f"Szukam tweetów zawierających '{search_term}' od {date_since}...")

tweets = pobierz_tweety(search_term, max_tweets, date_since)

if not tweets:
    print("Brak wyników.")
    exit()

df = pd.DataFrame(tweets, columns=["Data", "Użytkownik", "Treść", "Lajki", "Retweety", "Link"])
df['Data'] = pd.to_datetime(df['Data'])
df = df.sort_values(by='Data', ascending=False)

csv_filename = f"twitter_{search_term}.csv"
excel_filename = f"twitter_{search_term}.xlsx"

df.to_csv(csv_filename, index=False, encoding='utf-8')

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill

    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Tweetów')

        ws = writer.sheets['Tweetów']

        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="1DA1F2")  # Twitter blue

        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = fill
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                20 if column in ["Data", "Użytkownik", "Lajki", "Retweety"]
                else 70 if column == "Link"
                else 100
            )

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    print(f"\n✅ Zapisano dane do:")
    print(f"📄 CSV: {csv_filename}")
    print(f"📘 Excel: {excel_filename}")

except ImportError:
    print("Brak biblioteki openpyxl - zapisano tylko CSV")

# Podsumowanie
print("\n📊 Podsumowanie:")
print(f"🔹 Liczba tweetów: {len(df)}")
print(f"📅 Najstarszy tweet: {df['Data'].min().strftime('%Y-%m-%d')}")
print(f"📅 Najnowszy tweet: {df['Data'].max().strftime('%Y-%m-%d')}")
print(f"⬆️ Najwięcej lajków: {df['Lajki'].max()} ({df.loc[df['Lajki'].idxmax()]['Treść'][:50]}...)")
print(f"🔁 Najwięcej retweetów: {df['Retweety'].max()} ({df.loc[df['Retweety'].idxmax()]['Treść'][:50]}...)")
